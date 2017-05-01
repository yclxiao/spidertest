# -*- coding: UTF-8 -*-

import urllib2
from bs4 import BeautifulSoup
from openpyxl import Workbook

headers = [
    {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.133 Safari/537.36'}
]


def do_spider():
    url = 'https://www.douban.com/tag/%E5%AE%97%E6%95%99/book?start=120'

    req = urllib2.Request(url, headers=headers[0])

    source_code = urllib2.urlopen(req).read()

    plain_text = str(source_code)

    soup = BeautifulSoup(plain_text, "lxml")

    list_soup = soup.find('div', {'class': 'mod book-list'})

    list_soup_dd = list_soup.findAll('dd')

    book_infos = []

    for book_info in list_soup_dd:
        title_obj = book_info.find('a', {'class': 'title'})
        desc_obj = book_info.find('div', {'class': 'desc'})
        rate_num_obj = book_info.find('span', {'class': 'rating_nums'})

        if title_obj != None:
            title_info = title_obj.string.strip()
            title_href = title_obj.get('href')
        else:
            title_info = ''
            title_href = ''

        if desc_obj != None:
            desc_info = desc_obj.string.strip()
        else:
            desc_info = ''

        if rate_num_obj != None:
            rate_num_info = rate_num_obj.string.strip()
        else:
            rate_num_info = '0.0'

        print title_info + ' ' + title_href + ' ' + desc_info + ' ' + rate_num_info

        if len(desc_info) > 0:
            desces = desc_info.split('/')
            try:
                print '出版社：'.join(desces[-3:-2])
            except:
                print '出版社：暂无'

            try:
                print '出版日：'.join(desces[-2:-1])
            except:
                print '出版日：暂无'

            try:
                print '价格：'.join(desces[-1:])
            except:
                print '价格：暂无'

            try:
                print '作者：'.join(desces[0:1])
            except:
                print '作者：暂无'

        book_infos.append([title_info,title_href,desc_info,rate_num_info])

        book_infos = sorted(book_infos,key=lambda x:x[3],reverse=True)

    return book_infos


def spider_text_to_excel(book_infos):
    # for book_info in book_infos:
    #     print book_info[3]
    wb = Workbook()

    ws = []

    ws.append(wb.create_sheet("test",1))

    ws[0].append(['序号','书名','评分','作者','出版社'])

    count = 1
    for bl in book_infos:
        ws[0].append([count,bl[0],bl[3],bl[2],bl[1]])
        count += 1

    wb.save("book-list.xlsx")



if __name__ == '__main__':
    # 定义爬去的tag
    tag_list = ['个人管理', '时间管理']
    # 开始爬去
    book_infos = do_spider()
    # 将爬取的内容写到excel
    spider_text_to_excel(book_infos)
